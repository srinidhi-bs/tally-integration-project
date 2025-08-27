#!/usr/bin/env python3
"""
Professional Data Table Widget for TallyPrime Integration Manager

This module provides a comprehensive data table widget with advanced features:
- QTableView with professional styling and Qt MVC architecture
- Advanced sorting, filtering, and search capabilities  
- Context menus with export and refresh options
- Integration with theme manager for dark/light mode support
- Real-time data updates with caching integration
- Professional error handling and logging

Key Features:
- Multi-column sorting and filtering
- Context menu with export to CSV/Excel/PDF
- Keyboard navigation and shortcuts
- Professional data formatting and display
- Real-time search with highlighting
- Data validation and error display
- Integration with TallyPrime data reader and cache

Author: Srinidhi BS (Learning to code)
Assistant: Claude (Anthropic)
Date: August 27, 2025
Framework: PySide6 (Qt6)
"""

import logging
import csv
import json
from datetime import datetime
from typing import List, Optional, Dict, Any, Callable
from decimal import Decimal
from pathlib import Path

from PySide6.QtWidgets import (
    QWidget, QVBoxLayout, QHBoxLayout, QTableView, QLineEdit, QPushButton,
    QLabel, QComboBox, QHeaderView, QMenu, QApplication, QMessageBox,
    QFileDialog, QProgressDialog, QFrame, QSizePolicy, QToolButton,
    QCheckBox, QSpinBox, QDateEdit
)
from PySide6.QtCore import (
    Qt, QTimer, Signal, QThread, QSortFilterProxyModel, 
    QModelIndex, QPersistentModelIndex, QAbstractItemModel, QDate
)
from PySide6.QtGui import (
    QAction, QKeySequence, QShortcut, QIcon, QFont, QPalette,
    QStandardItemModel, QStandardItem
)

# Import our models and utilities
from core.models.ledger_model import LedgerInfo, LedgerTableModel, LedgerType
from ui.resources.styles.theme_manager import get_theme_manager, is_dark_theme
from core.utils.logger import get_logger

# Set up logger for this module
logger = get_logger(__name__)


class DataTableFilterProxyModel(QSortFilterProxyModel):
    """
    Professional proxy model for advanced filtering and sorting
    
    This proxy model provides:
    - Multi-column text filtering
    - Type-based filtering (ledger types)
    - Balance range filtering
    - Case-insensitive search
    - Professional sorting with custom logic
    
    Learning: QSortFilterProxyModel is Qt's standard approach for filtering
    and sorting without modifying the original model data.
    """
    
    def __init__(self, parent=None):
        """Initialize the filter proxy model"""
        super().__init__(parent)
        
        # Filter settings
        self._text_filter = ""
        self._type_filter = None
        self._min_balance = None
        self._max_balance = None
        self._show_zero_balances = True
        
        # Set case-insensitive filtering
        self.setFilterCaseSensitivity(Qt.CaseInsensitive)
        
        logger.debug("DataTableFilterProxyModel initialized")
    
    def set_text_filter(self, text: str):
        """Set text filter for searching across columns"""
        self._text_filter = text.lower().strip()
        self.invalidateFilter()
        logger.debug(f"Text filter set: '{text}'")
    
    def set_type_filter(self, ledger_type: Optional[LedgerType]):
        """Set filter by ledger type"""
        self._type_filter = ledger_type
        self.invalidateFilter()
        logger.debug(f"Type filter set: {ledger_type}")
    
    def set_balance_filter(self, min_balance: Optional[Decimal] = None, 
                          max_balance: Optional[Decimal] = None):
        """Set balance range filter"""
        self._min_balance = min_balance
        self._max_balance = max_balance
        self.invalidateFilter()
        logger.debug(f"Balance filter set: {min_balance} to {max_balance}")
    
    def set_show_zero_balances(self, show: bool):
        """Set whether to show ledgers with zero balance"""
        self._show_zero_balances = show
        self.invalidateFilter()
        logger.debug(f"Show zero balances: {show}")
    
    def filterAcceptsRow(self, source_row: int, source_parent: QModelIndex) -> bool:
        """
        Custom filtering logic for ledger data
        
        Args:
            source_row: Row index in source model
            source_parent: Parent index in source model
            
        Returns:
            bool: True if row should be visible
        """
        model = self.sourceModel()
        if not model or source_row >= model.rowCount():
            return False
        
        # Get ledger data from the model
        ledger_index = model.index(source_row, 0, source_parent)
        ledger = model.get_ledger(ledger_index)
        
        if not ledger:
            return False
        
        # Apply text filter (search in name, alias, group)
        if self._text_filter:
            searchable_text = " ".join([
                ledger.name.lower(),
                ledger.alias.lower(),
                ledger.parent_group_name.lower(),
                ledger.tax_info.gstin.lower()
            ])
            if self._text_filter not in searchable_text:
                return False
        
        # Apply type filter
        if self._type_filter and ledger.ledger_type != self._type_filter:
            return False
        
        # Apply balance filters
        balance = abs(ledger.balance.current_balance)
        
        if not self._show_zero_balances and balance == 0:
            return False
        
        if self._min_balance is not None and balance < self._min_balance:
            return False
        
        if self._max_balance is not None and balance > self._max_balance:
            return False
        
        return True
    
    def lessThan(self, source_left: QModelIndex, source_right: QModelIndex) -> bool:
        """
        Custom sorting logic for different column types
        
        Args:
            source_left: Left item index
            source_right: Right item index
            
        Returns:
            bool: True if left item should come before right item
        """
        model = self.sourceModel()
        if not model:
            return super().lessThan(source_left, source_right)
        
        column = source_left.column()
        
        # Get ledger data for both rows
        left_ledger = model.get_ledger(source_left)
        right_ledger = model.get_ledger(source_right)
        
        if not left_ledger or not right_ledger:
            return super().lessThan(source_left, source_right)
        
        # Custom sorting logic based on column
        if column == 0:  # Ledger Name
            return left_ledger.get_display_name().lower() < right_ledger.get_display_name().lower()
        elif column == 1:  # Group
            return left_ledger.parent_group_name.lower() < right_ledger.parent_group_name.lower()
        elif column == 2:  # Balance (sort by absolute value)
            return abs(left_ledger.balance.current_balance) < abs(right_ledger.balance.current_balance)
        elif column == 6:  # Voucher Count (numeric)
            return left_ledger.voucher_count < right_ledger.voucher_count
        else:
            # Default string comparison
            return super().lessThan(source_left, source_right)


class DataExportWorker(QThread):
    """
    Background worker thread for data export operations
    
    This worker handles:
    - CSV export with custom formatting
    - Excel export (if openpyxl available)
    - PDF export (if reportlab available)
    - Progress reporting during export
    - Error handling and reporting
    
    Learning: QThread is Qt's way to perform long-running operations
    without blocking the UI thread.
    """
    
    # Signals for communication with main thread
    progress_updated = Signal(int)  # Progress percentage (0-100)
    status_updated = Signal(str)    # Status message
    export_completed = Signal(str)  # Success message with file path
    export_failed = Signal(str)     # Error message
    
    def __init__(self, ledgers: List[LedgerInfo], file_path: str, export_format: str):
        """
        Initialize export worker
        
        Args:
            ledgers: List of ledger data to export
            file_path: Output file path
            export_format: Export format ('csv', 'excel', 'pdf')
        """
        super().__init__()
        self.ledgers = ledgers
        self.file_path = file_path
        self.export_format = export_format.lower()
        
        logger.info(f"Export worker initialized: {len(ledgers)} ledgers to {export_format}")
    
    def run(self):
        """Execute the export operation"""
        try:
            self.status_updated.emit("Starting export...")
            self.progress_updated.emit(0)
            
            if self.export_format == 'csv':
                self._export_csv()
            elif self.export_format == 'excel':
                self._export_excel()
            elif self.export_format == 'pdf':
                self._export_pdf()
            else:
                raise ValueError(f"Unsupported export format: {self.export_format}")
            
            self.progress_updated.emit(100)
            self.export_completed.emit(self.file_path)
            
        except Exception as e:
            logger.error(f"Export failed: {e}")
            self.export_failed.emit(str(e))
    
    def _export_csv(self):
        """Export data to CSV format"""
        self.status_updated.emit("Exporting to CSV...")
        
        headers = [
            "Ledger Name", "Alias", "Group", "Balance", "Balance Type", 
            "Ledger Type", "GST Number", "Contact Person", "Phone", "Email",
            "Last Voucher Date", "Voucher Count", "Credit Limit"
        ]
        
        with open(self.file_path, 'w', newline='', encoding='utf-8') as csvfile:
            writer = csv.writer(csvfile)
            writer.writerow(headers)
            
            total_ledgers = len(self.ledgers)
            for i, ledger in enumerate(self.ledgers):
                row = [
                    ledger.name,
                    ledger.alias,
                    ledger.parent_group_name,
                    float(ledger.balance.current_balance),
                    ledger.balance.balance_type.value,
                    ledger.ledger_type.value.replace('_', ' ').title(),
                    ledger.tax_info.gstin,
                    ledger.contact_info.contact_person,
                    ledger.contact_info.phone,
                    ledger.contact_info.email,
                    ledger.last_voucher_date.isoformat() if ledger.last_voucher_date else "",
                    ledger.voucher_count,
                    float(ledger.credit_limit)
                ]
                writer.writerow(row)
                
                # Update progress
                progress = int((i + 1) / total_ledgers * 100)
                self.progress_updated.emit(progress)
        
        logger.info(f"CSV export completed: {self.file_path}")
    
    def _export_excel(self):
        """Export data to Excel format (requires openpyxl)"""
        try:
            import openpyxl
            from openpyxl.styles import Font, PatternFill
        except ImportError:
            raise ImportError("openpyxl package required for Excel export")
        
        self.status_updated.emit("Exporting to Excel...")
        
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Ledger Data"
        
        # Headers with styling
        headers = [
            "Ledger Name", "Alias", "Group", "Balance", "Balance Type", 
            "Ledger Type", "GST Number", "Contact Person", "Phone", "Email",
            "Last Voucher Date", "Voucher Count", "Credit Limit"
        ]
        
        # Style headers
        header_font = Font(bold=True)
        header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
        
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col, value=header)
            cell.font = header_font
            cell.fill = header_fill
        
        # Data rows
        total_ledgers = len(self.ledgers)
        for i, ledger in enumerate(self.ledgers):
            row = i + 2  # Start from row 2 (after headers)
            
            ws.cell(row=row, column=1, value=ledger.name)
            ws.cell(row=row, column=2, value=ledger.alias)
            ws.cell(row=row, column=3, value=ledger.parent_group_name)
            ws.cell(row=row, column=4, value=float(ledger.balance.current_balance))
            ws.cell(row=row, column=5, value=ledger.balance.balance_type.value)
            ws.cell(row=row, column=6, value=ledger.ledger_type.value.replace('_', ' ').title())
            ws.cell(row=row, column=7, value=ledger.tax_info.gstin)
            ws.cell(row=row, column=8, value=ledger.contact_info.contact_person)
            ws.cell(row=row, column=9, value=ledger.contact_info.phone)
            ws.cell(row=row, column=10, value=ledger.contact_info.email)
            ws.cell(row=row, column=11, value=ledger.last_voucher_date.isoformat() if ledger.last_voucher_date else "")
            ws.cell(row=row, column=12, value=ledger.voucher_count)
            ws.cell(row=row, column=13, value=float(ledger.credit_limit))
            
            # Update progress
            progress = int((i + 1) / total_ledgers * 100)
            self.progress_updated.emit(progress)
        
        # Auto-adjust column widths
        for column in ws.columns:
            max_length = 0
            column_letter = column[0].column_letter
            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            adjusted_width = (max_length + 2) * 1.2
            ws.column_dimensions[column_letter].width = adjusted_width
        
        wb.save(self.file_path)
        logger.info(f"Excel export completed: {self.file_path}")
    
    def _export_pdf(self):
        """Export data to PDF format (requires reportlab)"""
        try:
            from reportlab.lib.pagesizes import letter, A4
            from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph
            from reportlab.lib.styles import getSampleStyleSheet
            from reportlab.lib.units import inch
            from reportlab.lib import colors
        except ImportError:
            raise ImportError("reportlab package required for PDF export")
        
        self.status_updated.emit("Exporting to PDF...")
        
        doc = SimpleDocTemplate(self.file_path, pagesize=A4)
        story = []
        
        # Title
        styles = getSampleStyleSheet()
        title = Paragraph("TallyPrime Ledger Report", styles['Title'])
        story.append(title)
        
        # Prepare data for table
        headers = ["Name", "Group", "Balance", "Type", "GST No.", "Vouchers"]
        data = [headers]
        
        total_ledgers = len(self.ledgers)
        for i, ledger in enumerate(self.ledgers):
            row = [
                ledger.get_display_name()[:20],  # Truncate for PDF
                ledger.parent_group_name[:15],
                ledger.get_balance_display(),
                ledger.ledger_type.value.replace('_', ' ').title()[:10],
                ledger.tax_info.gstin,
                str(ledger.voucher_count)
            ]
            data.append(row)
            
            # Update progress
            progress = int((i + 1) / total_ledgers * 100)
            self.progress_updated.emit(progress)
        
        # Create table with styling
        table = Table(data)
        table.setStyle(TableStyle([
            ('BACKGROUND', (0, 0), (-1, 0), colors.grey),
            ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
            ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
            ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
            ('FONTSIZE', (0, 0), (-1, 0), 10),
            ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
            ('BACKGROUND', (0, 1), (-1, -1), colors.beige),
            ('FONTNAME', (0, 1), (-1, -1), 'Helvetica'),
            ('FONTSIZE', (0, 1), (-1, -1), 8),
            ('GRID', (0, 0), (-1, -1), 1, colors.black)
        ]))
        
        story.append(table)
        doc.build(story)
        
        logger.info(f"PDF export completed: {self.file_path}")


class ProfessionalDataTableWidget(QWidget):
    """
    Professional Data Table Widget for TallyPrime Integration Manager
    
    This comprehensive widget provides:
    - Advanced table display with Qt MVC architecture
    - Professional filtering and search capabilities
    - Context menu with export options
    - Keyboard shortcuts and navigation
    - Theme-aware styling (dark/light mode support)
    - Real-time data updates and caching integration
    - Professional error handling and user feedback
    
    Signals:
        ledger_selected: Emitted when a ledger is selected (ledger: LedgerInfo)
        refresh_requested: Emitted when refresh is requested
        filter_changed: Emitted when filter settings change
    """
    
    # Signals for external communication
    ledger_selected = Signal(object)  # LedgerInfo object
    refresh_requested = Signal()
    filter_changed = Signal(dict)  # Filter settings dictionary
    
    def __init__(self, parent=None):
        """
        Initialize the professional data table widget
        
        Args:
            parent: Parent widget (optional)
        """
        super().__init__(parent)
        
        # Initialize components
        self.ledger_model = LedgerTableModel()
        self.filter_model = DataTableFilterProxyModel(self)
        self.export_worker = None
        self.theme_manager = get_theme_manager()
        
        # Setup the user interface
        self._setup_ui()
        self._setup_models()
        self._setup_connections()
        self._setup_shortcuts()
        self._apply_theme()
        
        # Connect to theme manager for dynamic theming
        self.theme_manager.theme_changed.connect(self._apply_theme)
        
        logger.info("ProfessionalDataTableWidget initialized")
    
    def _setup_ui(self):
        """Setup the user interface components"""
        # Main layout
        layout = QVBoxLayout()
        layout.setContentsMargins(8, 8, 8, 8)
        layout.setSpacing(8)
        
        # Top toolbar with filters
        toolbar_frame = QFrame()
        toolbar_frame.setFrameStyle(QFrame.StyledPanel)
        toolbar_layout = QHBoxLayout(toolbar_frame)
        toolbar_layout.setContentsMargins(8, 6, 8, 6)
        
        # Search box
        search_label = QLabel("🔍 Search:")
        self.search_edit = QLineEdit()
        self.search_edit.setPlaceholderText("Search ledgers by name, group, or GST number...")
        self.search_edit.setMinimumWidth(300)
        
        # Type filter combo
        type_label = QLabel("Type:")
        self.type_combo = QComboBox()
        self.type_combo.addItem("All Types", None)
        for ledger_type in LedgerType:
            self.type_combo.addItem(ledger_type.value.replace('_', ' ').title(), ledger_type)
        self.type_combo.setMinimumWidth(150)
        
        # Balance filter
        balance_label = QLabel("Balance:")
        self.balance_min_edit = QLineEdit()
        self.balance_min_edit.setPlaceholderText("Min")
        self.balance_min_edit.setMaximumWidth(80)
        
        balance_to_label = QLabel("to")
        self.balance_max_edit = QLineEdit()
        self.balance_max_edit.setPlaceholderText("Max")
        self.balance_max_edit.setMaximumWidth(80)
        
        # Show zero balances checkbox
        self.show_zero_check = QCheckBox("Show Zero Balances")
        self.show_zero_check.setChecked(True)
        
        # Action buttons
        self.refresh_btn = QPushButton("🔄 Refresh")
        self.refresh_btn.setToolTip("Refresh data from TallyPrime")
        
        self.export_btn = QPushButton("📊 Export")
        self.export_btn.setToolTip("Export data to file")
        
        self.clear_filter_btn = QPushButton("✖ Clear Filters")
        self.clear_filter_btn.setToolTip("Clear all filters")
        
        # Add toolbar components
        toolbar_layout.addWidget(search_label)
        toolbar_layout.addWidget(self.search_edit)
        toolbar_layout.addSpacing(10)
        toolbar_layout.addWidget(type_label)
        toolbar_layout.addWidget(self.type_combo)
        toolbar_layout.addSpacing(10)
        toolbar_layout.addWidget(balance_label)
        toolbar_layout.addWidget(self.balance_min_edit)
        toolbar_layout.addWidget(balance_to_label)
        toolbar_layout.addWidget(self.balance_max_edit)
        toolbar_layout.addSpacing(10)
        toolbar_layout.addWidget(self.show_zero_check)
        toolbar_layout.addStretch()
        toolbar_layout.addWidget(self.clear_filter_btn)
        toolbar_layout.addWidget(self.refresh_btn)
        toolbar_layout.addWidget(self.export_btn)
        
        layout.addWidget(toolbar_frame)
        
        # Table view
        self.table_view = QTableView()
        self.table_view.setAlternatingRowColors(True)
        self.table_view.setSelectionBehavior(QTableView.SelectRows)
        self.table_view.setSelectionMode(QTableView.SingleSelection)
        self.table_view.setSortingEnabled(True)
        self.table_view.setContextMenuPolicy(Qt.CustomContextMenu)
        
        # Configure headers
        self.table_view.horizontalHeader().setStretchLastSection(True)
        self.table_view.horizontalHeader().setSectionResizeMode(QHeaderView.Interactive)
        self.table_view.verticalHeader().setVisible(False)
        
        layout.addWidget(self.table_view)
        
        # Status bar
        status_frame = QFrame()
        status_frame.setFrameStyle(QFrame.StyledPanel)
        status_layout = QHBoxLayout(status_frame)
        status_layout.setContentsMargins(8, 4, 8, 4)
        
        self.status_label = QLabel("Ready")
        self.count_label = QLabel("0 ledgers")
        
        status_layout.addWidget(self.status_label)
        status_layout.addStretch()
        status_layout.addWidget(self.count_label)
        
        layout.addWidget(status_frame)
        
        self.setLayout(layout)
        
        # Set initial size
        self.resize(1000, 600)
    
    def _setup_models(self):
        """Setup the data models and proxy model"""
        # Set up filter proxy model
        self.filter_model.setSourceModel(self.ledger_model)
        self.table_view.setModel(self.filter_model)
        
        # Set default sorting
        self.filter_model.sort(0, Qt.AscendingOrder)
        
        logger.debug("Models configured successfully")
    
    def _setup_connections(self):
        """Setup signal-slot connections"""
        # Filter connections
        self.search_edit.textChanged.connect(self._on_search_changed)
        self.type_combo.currentIndexChanged.connect(self._on_type_filter_changed)
        self.balance_min_edit.textChanged.connect(self._on_balance_filter_changed)
        self.balance_max_edit.textChanged.connect(self._on_balance_filter_changed)
        self.show_zero_check.toggled.connect(self._on_show_zero_changed)
        
        # Button connections
        self.refresh_btn.clicked.connect(self._on_refresh_clicked)
        self.export_btn.clicked.connect(self._on_export_clicked)
        self.clear_filter_btn.clicked.connect(self._clear_all_filters)
        
        # Table connections
        self.table_view.doubleClicked.connect(self._on_row_double_clicked)
        self.table_view.customContextMenuRequested.connect(self._show_context_menu)
        
        # Model connections
        self.ledger_model.modelReset.connect(self._update_status)
        
        logger.debug("Signal connections established")
    
    def _setup_shortcuts(self):
        """Setup keyboard shortcuts"""
        # Refresh shortcut (F5)
        refresh_shortcut = QShortcut(QKeySequence("F5"), self)
        refresh_shortcut.activated.connect(self._on_refresh_clicked)
        
        # Search shortcut (Ctrl+F)
        search_shortcut = QShortcut(QKeySequence("Ctrl+F"), self)
        search_shortcut.activated.connect(self.search_edit.setFocus)
        
        # Export shortcut (Ctrl+E)
        export_shortcut = QShortcut(QKeySequence("Ctrl+E"), self)
        export_shortcut.activated.connect(self._on_export_clicked)
        
        # Clear filters shortcut (Ctrl+R)
        clear_shortcut = QShortcut(QKeySequence("Ctrl+R"), self)
        clear_shortcut.activated.connect(self._clear_all_filters)
        
        logger.debug("Keyboard shortcuts configured")
    
    def _apply_theme(self):
        """Apply current theme styling"""
        try:
            # Get theme colors
            colors = self.theme_manager.colors
            
            # Generate custom stylesheet for data table
            stylesheet = f"""
                QFrame {{
                    background-color: {colors['surface']};
                    border: 1px solid {colors['border']};
                    border-radius: 6px;
                }}
                
                QTableView {{
                    background-color: {colors['surface']};
                    color: {colors['text_primary']};
                    gridline-color: {colors['border_light']};
                    selection-background-color: {colors['selection']};
                    selection-color: {colors['selection_text']};
                    alternate-background-color: {colors['surface_variant']};
                    font-family: "Segoe UI", Arial, sans-serif;
                    font-size: 12px;
                }}
                
                QHeaderView::section {{
                    background-color: {colors['secondary']};
                    color: white;
                    padding: 8px;
                    border: 1px solid {colors['border']};
                    font-weight: bold;
                    font-size: 11px;
                }}
                
                QLineEdit {{
                    border: 2px solid {colors['border_light']};
                    border-radius: 6px;
                    padding: 6px 10px;
                    font-size: 12px;
                    background-color: {colors['surface']};
                    color: {colors['text_primary']};
                }}
                
                QLineEdit:focus {{
                    border-color: {colors['primary']};
                }}
                
                QComboBox {{
                    border: 2px solid {colors['border_light']};
                    border-radius: 6px;
                    padding: 6px 10px;
                    font-size: 12px;
                    background-color: {colors['surface']};
                    color: {colors['text_primary']};
                }}
                
                QPushButton {{
                    background-color: {colors['primary']};
                    border: none;
                    color: white;
                    padding: 8px 16px;
                    border-radius: 6px;
                    font-weight: bold;
                    font-size: 12px;
                }}
                
                QPushButton:hover {{
                    background-color: {colors['primary_hover']};
                }}
                
                QPushButton:pressed {{
                    background-color: {colors['primary_pressed']};
                }}
                
                QCheckBox {{
                    font-size: 12px;
                    spacing: 6px;
                    color: {colors['text_primary']};
                }}
                
                QLabel {{
                    color: {colors['text_secondary']};
                    font-size: 12px;
                    font-weight: bold;
                }}
            """
            
            self.setStyleSheet(stylesheet)
            logger.debug("Theme applied successfully")
            
        except Exception as e:
            logger.error(f"Failed to apply theme: {e}")
    
    # Event handlers (shortened for space - continuing in next message)
    def _on_search_changed(self):
        """Handle search text changes"""
        QTimer.singleShot(300, self._apply_filters)  # Debounce search
    
    def _on_type_filter_changed(self):
        """Handle type filter changes"""
        self._apply_filters()
    
    def _on_balance_filter_changed(self):
        """Handle balance filter changes"""
        QTimer.singleShot(500, self._apply_filters)  # Debounce balance input
    
    def _on_show_zero_changed(self):
        """Handle show zero balances toggle"""
        self._apply_filters()
    
    def _apply_filters(self):
        """Apply all current filter settings"""
        try:
            # Text filter
            search_text = self.search_edit.text().strip()
            self.filter_model.set_text_filter(search_text)
            
            # Type filter
            type_data = self.type_combo.currentData()
            self.filter_model.set_type_filter(type_data)
            
            # Balance filter
            min_balance = None
            max_balance = None
            
            try:
                min_text = self.balance_min_edit.text().strip()
                if min_text:
                    min_balance = Decimal(min_text)
            except (ValueError, ArithmeticError):
                pass
            
            try:
                max_text = self.balance_max_edit.text().strip()
                if max_text:
                    max_balance = Decimal(max_text)
            except (ValueError, ArithmeticError):
                pass
            
            self.filter_model.set_balance_filter(min_balance, max_balance)
            
            # Show zero balances
            self.filter_model.set_show_zero_balances(self.show_zero_check.isChecked())
            
            # Update status
            self._update_status()
            
            # Emit filter changed signal
            filter_settings = {
                'search_text': search_text,
                'type_filter': type_data,
                'min_balance': min_balance,
                'max_balance': max_balance,
                'show_zero_balances': self.show_zero_check.isChecked()
            }
            self.filter_changed.emit(filter_settings)
            
            logger.debug("Filters applied successfully")
            
        except Exception as e:
            logger.error(f"Error applying filters: {e}")
            self.status_label.setText(f"Filter error: {e}")
    
    def _clear_all_filters(self):
        """Clear all filter settings"""
        self.search_edit.clear()
        self.type_combo.setCurrentIndex(0)
        self.balance_min_edit.clear()
        self.balance_max_edit.clear()
        self.show_zero_check.setChecked(True)
        
        # Reapply filters
        self._apply_filters()
        
        self.status_label.setText("Filters cleared")
        logger.info("All filters cleared")
    
    def _on_refresh_clicked(self):
        """Handle refresh button click"""
        self.refresh_requested.emit()
        self.status_label.setText("Refreshing data...")
        logger.info("Refresh requested")
    
    def _on_export_clicked(self):
        """Handle export button click"""
        try:
            # Get visible (filtered) ledgers
            visible_ledgers = []
            for row in range(self.filter_model.rowCount()):
                source_index = self.filter_model.mapToSource(self.filter_model.index(row, 0))
                ledger = self.ledger_model.get_ledger(source_index)
                if ledger:
                    visible_ledgers.append(ledger)
            
            if not visible_ledgers:
                QMessageBox.information(self, "No Data", "No ledgers to export.")
                return
            
            # Show export options
            self._show_export_dialog(visible_ledgers)
            
        except Exception as e:
            logger.error(f"Export preparation failed: {e}")
            QMessageBox.critical(self, "Export Error", f"Failed to prepare export: {e}")
    
    def _show_export_dialog(self, ledgers: List[LedgerInfo]):
        """Show export format selection dialog"""
        formats = {
            "CSV File (*.csv)": "csv",
            "Excel File (*.xlsx)": "excel", 
            "PDF Report (*.pdf)": "pdf"
        }
        
        format_filter = ";;".join(formats.keys())
        
        file_path, selected_filter = QFileDialog.getSaveFileName(
            self,
            "Export Ledger Data",
            f"ledger_data_{datetime.now().strftime('%Y%m%d_%H%M%S')}.csv",
            format_filter
        )
        
        if file_path:
            export_format = formats.get(selected_filter, "csv")
            self._start_export(ledgers, file_path, export_format)
    
    def _start_export(self, ledgers: List[LedgerInfo], file_path: str, export_format: str):
        """Start background export operation"""
        try:
            # Create and configure export worker
            self.export_worker = DataExportWorker(ledgers, file_path, export_format)
            
            # Create progress dialog
            self.export_progress = QProgressDialog("Exporting data...", "Cancel", 0, 100, self)
            self.export_progress.setWindowModality(Qt.WindowModal)
            self.export_progress.setAutoClose(True)
            
            # Connect worker signals
            self.export_worker.progress_updated.connect(self.export_progress.setValue)
            self.export_worker.status_updated.connect(self.export_progress.setLabelText)
            self.export_worker.export_completed.connect(self._on_export_completed)
            self.export_worker.export_failed.connect(self._on_export_failed)
            self.export_progress.canceled.connect(self.export_worker.terminate)
            
            # Start export
            self.export_worker.start()
            self.export_progress.show()
            
            logger.info(f"Export started: {len(ledgers)} ledgers to {export_format}")
            
        except Exception as e:
            logger.error(f"Failed to start export: {e}")
            QMessageBox.critical(self, "Export Error", f"Failed to start export: {e}")
    
    def _on_export_completed(self, file_path: str):
        """Handle successful export completion"""
        self.export_progress.close()
        self.status_label.setText(f"Export completed: {Path(file_path).name}")
        
        QMessageBox.information(
            self, 
            "Export Successful", 
            f"Data exported successfully to:\n{file_path}"
        )
        
        logger.info(f"Export completed successfully: {file_path}")
    
    def _on_export_failed(self, error_message: str):
        """Handle export failure"""
        self.export_progress.close()
        self.status_label.setText("Export failed")
        
        QMessageBox.critical(self, "Export Failed", f"Export failed:\n{error_message}")
        
        logger.error(f"Export failed: {error_message}")
    
    def _on_row_double_clicked(self, index: QModelIndex):
        """Handle double-click on table row"""
        if index.isValid():
            # Map to source model
            source_index = self.filter_model.mapToSource(index)
            ledger = self.ledger_model.get_ledger(source_index)
            
            if ledger:
                self.ledger_selected.emit(ledger)
                logger.debug(f"Ledger selected: {ledger.name}")
    
    def _show_context_menu(self, position):
        """Show context menu for table"""
        index = self.table_view.indexAt(position)
        if not index.isValid():
            return
        
        # Create context menu
        menu = QMenu(self)
        
        # Get ledger info for context
        source_index = self.filter_model.mapToSource(index)
        ledger = self.ledger_model.get_ledger(source_index)
        
        if ledger:
            # View details action
            view_action = QAction(f"📋 View Details: {ledger.name}", self)
            view_action.triggered.connect(lambda: self.ledger_selected.emit(ledger))
            menu.addAction(view_action)
            
            menu.addSeparator()
            
            # Copy actions
            copy_name_action = QAction("📋 Copy Name", self)
            copy_name_action.triggered.connect(
                lambda: QApplication.clipboard().setText(ledger.name)
            )
            menu.addAction(copy_name_action)
            
            if ledger.tax_info.gstin:
                copy_gstin_action = QAction("📋 Copy GST Number", self)
                copy_gstin_action.triggered.connect(
                    lambda: QApplication.clipboard().setText(ledger.tax_info.gstin)
                )
                menu.addAction(copy_gstin_action)
            
            menu.addSeparator()
        
        # Export actions
        export_all_action = QAction("📊 Export All Data", self)
        export_all_action.triggered.connect(self._on_export_clicked)
        menu.addAction(export_all_action)
        
        export_filtered_action = QAction("📊 Export Filtered Data", self)
        export_filtered_action.triggered.connect(self._on_export_clicked)
        menu.addAction(export_filtered_action)
        
        menu.addSeparator()
        
        # Refresh action
        refresh_action = QAction("🔄 Refresh Data", self)
        refresh_action.triggered.connect(self._on_refresh_clicked)
        menu.addAction(refresh_action)
        
        # Show menu
        menu.exec_(self.table_view.mapToGlobal(position))
    
    def _update_status(self):
        """Update status bar with current data counts"""
        total_count = self.ledger_model.rowCount()
        filtered_count = self.filter_model.rowCount()
        
        if total_count == filtered_count:
            self.count_label.setText(f"{total_count} ledgers")
        else:
            self.count_label.setText(f"{filtered_count} of {total_count} ledgers")
        
        if filtered_count == 0 and total_count > 0:
            self.status_label.setText("No ledgers match current filters")
        elif total_count == 0:
            self.status_label.setText("No data loaded")
        else:
            self.status_label.setText("Ready")
    
    # Public API methods for integration
    
    def set_ledger_data(self, ledgers: List[LedgerInfo]):
        """
        Set ledger data to display in the table
        
        Args:
            ledgers: List of LedgerInfo instances
        """
        try:
            self.ledger_model.update_ledgers(ledgers)
            self._update_status()
            self.status_label.setText("Data loaded successfully")
            logger.info(f"Data table updated with {len(ledgers)} ledgers")
            
        except Exception as e:
            logger.error(f"Failed to set ledger data: {e}")
            self.status_label.setText("Error loading data")
    
    def get_selected_ledger(self) -> Optional[LedgerInfo]:
        """
        Get currently selected ledger
        
        Returns:
            LedgerInfo: Selected ledger or None
        """
        selection = self.table_view.selectionModel()
        if selection.hasSelection():
            index = selection.currentIndex()
            source_index = self.filter_model.mapToSource(index)
            return self.ledger_model.get_ledger(source_index)
        return None
    
    def select_ledger_by_name(self, ledger_name: str) -> bool:
        """
        Select ledger by name
        
        Args:
            ledger_name: Name of ledger to select
            
        Returns:
            bool: True if ledger found and selected
        """
        for row in range(self.filter_model.rowCount()):
            index = self.filter_model.index(row, 0)
            source_index = self.filter_model.mapToSource(index)
            ledger = self.ledger_model.get_ledger(source_index)
            
            if ledger and ledger.name == ledger_name:
                self.table_view.selectRow(row)
                self.table_view.scrollTo(index)
                return True
        
        return False
    
    def get_filter_settings(self) -> Dict[str, Any]:
        """
        Get current filter settings
        
        Returns:
            Dict containing current filter settings
        """
        return {
            'search_text': self.search_edit.text(),
            'type_filter': self.type_combo.currentData(),
            'min_balance': self.balance_min_edit.text(),
            'max_balance': self.balance_max_edit.text(),
            'show_zero_balances': self.show_zero_check.isChecked()
        }
    
    def apply_filter_settings(self, settings: Dict[str, Any]):
        """
        Apply filter settings
        
        Args:
            settings: Dictionary containing filter settings
        """
        if 'search_text' in settings:
            self.search_edit.setText(settings['search_text'])
        
        if 'type_filter' in settings:
            type_filter = settings['type_filter']
            for i in range(self.type_combo.count()):
                if self.type_combo.itemData(i) == type_filter:
                    self.type_combo.setCurrentIndex(i)
                    break
        
        if 'min_balance' in settings:
            self.balance_min_edit.setText(str(settings['min_balance'] or ''))
        
        if 'max_balance' in settings:
            self.balance_max_edit.setText(str(settings['max_balance'] or ''))
        
        if 'show_zero_balances' in settings:
            self.show_zero_check.setChecked(settings['show_zero_balances'])
        
        self._apply_filters()
    
    def clear_selection(self):
        """Clear table selection"""
        self.table_view.clearSelection()
    
    def refresh_display(self):
        """Refresh the display without reloading data"""
        self.filter_model.invalidate()
        self._update_status()
        logger.debug("Display refreshed")
    
    def get_visible_ledger_count(self) -> int:
        """Get count of visible (filtered) ledgers"""
        return self.filter_model.rowCount()
    
    def get_total_ledger_count(self) -> int:
        """Get total count of all ledgers"""
        return self.ledger_model.rowCount()